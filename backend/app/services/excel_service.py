import io
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── helpers ────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="FF000000", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

DARK_NAVY   = "FF1A1A2E"
MID_NAVY    = "FF16213E"
GREEN_ACC   = "FF27AE60"
RED_ACC     = "FFE74C3C"
BLUE_ACC    = "FF2980B9"
LIGHT_BG    = "FFF5F5F5"
WHITE       = "FFFFFFFF"
CARD_BG     = "FFF9F9F9"
SECTION_BG  = "FFE8E8F0"
WHITE_FONT  = "FFFFFFFF"
GREEN_FONT  = "FFE74C3C"   # template uses red for Expense text
RED_FONT    = "FFE74C3C"
BLUE_FONT   = "FF2980B9"

RUPIAH_FMT  = '"Rp" #,##0'


def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ── main function ───────────────────────────────────────────────────────────

def generate_excel_report(data, month, year):
    """
    Generate a Kalren financial report Excel workbook.

    Parameters
    ----------
    data  : list[dict]  – each dict with keys matching the original code:
                          tanggal, flow, kategori, sub_kategori,
                          keterangan, metode, gross_amount, potongan, net_amount
    month : str/int     – month label (e.g. "Mei" or 5)
    year  : str/int     – year  label (e.g. 2026)
    """

    # ── normalise rows ──────────────────────────────────────────────────────
    rows = []
    for item in data:
        rows.append({
            "Tanggal":      item.get("tanggal"),
            "Flow":         item.get("flow"),
            "Kategori":     item.get("kategori"),
            "Sub Kategori": item.get("sub_kategori") or "-",
            "Keterangan":   item.get("keterangan"),
            "Metode":       item.get("metode"),
            "Gross Amount": item.get("gross_amount", 0),
            "Potongan":     item.get("potongan", 0),
            "Net Amount":   item.get("net_amount", 0),
        })

    # ── aggregate ───────────────────────────────────────────────────────────
    total_income  = sum(r["Net Amount"] for r in rows if r["Flow"] == "Income")
    total_expense = sum(r["Net Amount"] for r in rows if r["Flow"] == "Expense")
    net_profit    = total_income - total_expense

    # per-kategori totals  {kategori: {"income": x, "expense": y}}
    kat_totals = defaultdict(lambda: {"income": 0, "expense": 0})
    for r in rows:
        k = r["Kategori"] or "Lainnya"
        if r["Flow"] == "Income":
            kat_totals[k]["income"] += r["Net Amount"]
        else:
            kat_totals[k]["expense"] += r["Net Amount"]

    # income breakdown by sub-kategori
    income_items   = defaultdict(float)
    expense_items  = defaultdict(float)
    for r in rows:
        label = f"{r['Kategori']} - {r['Sub Kategori']}"
        if r["Flow"] == "Income":
            income_items[r["Sub Kategori"] or r["Kategori"]] += r["Net Amount"]
        else:
            expense_items[label] += r["Net Amount"]

    period_label = f"{month} {year}"

    # ── workbook ────────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)     # remove default sheet

    _build_cover(wb, period_label)
    _build_dashboard(wb, period_label, total_income, total_expense, net_profit, kat_totals)
    _build_detail(wb, period_label, rows)
    _build_ringkasan(wb, total_income, total_expense, net_profit, income_items, expense_items)
    _build_arus_kas(wb, period_label, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Cover ───────────────────────────────────────────────────────────────────

def _build_cover(wb, period_label):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # col widths
    for col in "ABCDEFGH":
        _set_col_width(ws, col, 14)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["H"].width = 3

    # row heights
    for r in range(1, 45):
        ws.row_dimensions[r].height = 18

    # ── full dark background rows 1–44 ──
    for r in range(1, 45):
        for c in range(1, 9):
            ws.cell(r, c).fill = _fill(DARK_NAVY)

    # ── accent bar left edge rows 8–38 ──
    for r in range(8, 39):
        ws.cell(r, 1).fill = _fill(GREEN_ACC)

    # ── logo (if exists)  B3:C7 ──
    logo_path = "app/assets/logokalren.png"
    if os.path.exists(logo_path):
        img = __import__("openpyxl.drawing.image", fromlist=["Image"]).Image(logo_path)
        img.width  = 80
        img.height = 80
        img.anchor = "B3"
        ws.add_image(img)

    # ── KALREN  B10 ──
    c = ws["B10"]
    c.value     = "KALREN"
    c.font      = _font(bold=True, size=28, color=WHITE_FONT)
    c.alignment = _align("left", "center")
    ws.row_dimensions[10].height = 36

    # ── thin green divider row 13 ──
    for col in range(2, 8):
        ws.cell(13, col).fill = _fill(GREEN_ACC)
    ws.row_dimensions[13].height = 3

    # ── LAPORAN KEUANGAN  B15 ──
    c = ws["B15"]
    c.value     = "LAPORAN KEUANGAN"
    c.font      = _font(size=14, color="FFB0B8C8")
    c.alignment = _align("left", "center")
    ws.row_dimensions[15].height = 24

    # ── Periode  B17 ──
    c = ws["B17"]
    c.value     = f"Periode: {period_label}"
    c.font      = _font(bold=True, size=13, color=WHITE_FONT)
    c.alignment = _align("left", "center")
    ws.row_dimensions[17].height = 22

    # ── thin green divider row 19 ──
    for col in range(2, 8):
        ws.cell(19, col).fill = _fill(GREEN_ACC)
    ws.row_dimensions[19].height = 3

    # ── footer area: slightly lighter bg rows 36–44 ──
    for r in range(36, 45):
        for col in range(1, 9):
            ws.cell(r, col).fill = _fill(MID_NAVY)

    # ── thin accent line row 36 ──
    for col in range(2, 8):
        ws.cell(36, col).fill = _fill(GREEN_ACC)
    ws.row_dimensions[36].height = 2

    # ── footer text  B39 ──
    ws.merge_cells("B39:G39")
    f = ws["B39"]
    f.value     = "Konfidensial · Untuk Internal Perusahaan"
    f.font      = _font(size=9, color="FF6C7A8D")
    f.alignment = _align("center", "center")


# ── Dashboard ───────────────────────────────────────────────────────────────

def _build_dashboard(wb, period_label, total_income, total_expense, net_profit, kat_totals):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    for col, w in [("A",2),("B",22),("C",22),("D",22),("E",22),("F",14),("G",14)]:
        _set_col_width(ws, col, w)

    # title bar  A2:G2
    for c in range(1, 8):
        ws.cell(2, c).fill = _fill(DARK_NAVY)
    ws.merge_cells("A2:G2")
    t = ws["A2"]
    t.value     = f"📊  DASHBOARD KEUANGAN — KALREN"
    t.font      = _font(bold=True, size=15, color=WHITE_FONT)
    t.alignment = _align("left")

    # KPI colour bars  row 5
    ws.cell(2, 2).fill = _fill(DARK_NAVY)
    for col, color in [(2, GREEN_ACC), (3, RED_ACC), (4, BLUE_ACC)]:
        ws.cell(5, col).fill = _fill(color)
        ws.row_dimensions[5].height = 6

    # KPI labels  row 6
    kpi_labels  = ["TOTAL PENDAPATAN", "TOTAL PENGELUARAN", "LABA BERSIH"]
    kpi_values  = [total_income, total_expense, net_profit]
    kpi_colors  = [GREEN_ACC, RED_ACC, BLUE_ACC]

    for i, (lbl, val, clr) in enumerate(zip(kpi_labels, kpi_values, kpi_colors)):
        col = 2 + i
        lc  = ws.cell(6, col)
        lc.value     = lbl
        lc.font      = _font(size=8, color="FF555555")
        lc.fill      = _fill(CARD_BG)
        lc.alignment = _align("left")

        vc  = ws.cell(7, col)
        vc.value          = val
        vc.number_format  = RUPIAH_FMT
        vc.font           = _font(bold=True, size=14, color=clr)
        vc.fill           = _fill(CARD_BG)

        for r in [8, 9]:
            ws.cell(r, col).fill = _fill(CARD_BG)

    # category table header  row 11-12
    ws.merge_cells("B11:E11")
    hdr = ws["B11"]
    hdr.value     = "Ringkasan Berdasarkan Kategori"
    hdr.font      = _font(bold=True, size=11)
    hdr.fill      = _fill(SECTION_BG)
    hdr.alignment = _align("left")

    col_headers = ["Kategori", "Pendapatan (Rp)", "Pengeluaran (Rp)", "Net (Rp)"]
    for j, h in enumerate(col_headers):
        c = ws.cell(12, 2 + j)
        c.value     = h
        c.font      = _font(bold=True, size=9, color=WHITE_FONT)
        c.fill      = _fill(DARK_NAVY)
        c.alignment = _align("center")

    # category rows
    for i, (kat, totals) in enumerate(kat_totals.items()):
        r   = 13 + i
        bg  = LIGHT_BG if i % 2 == 0 else WHITE
        inc = totals["income"]
        exp = totals["expense"]
        net = inc - exp

        vals = [kat, inc, exp, net]
        for j, v in enumerate(vals):
            c = ws.cell(r, 2 + j)
            c.value = v
            c.fill  = _fill(bg)
            c.font  = _font(size=9)
            if j > 0:
                c.number_format = RUPIAH_FMT
                c.font = _font(
                    size=9,
                    color=GREEN_ACC if (j == 3 and net >= 0) else RED_ACC if j == 3 else "FF000000"
                )


# ── Detail ───────────────────────────────────────────────────────────────────

def _build_detail(wb, period_label, rows):
    ws = wb.create_sheet("Detail")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    for col, w in [("A",2),("B",14),("C",10),("D",18),("E",18),
                   ("F",22),("G",12),("H",18),("I",16),("J",16),("K",2)]:
        _set_col_width(ws, col, w)

    # title  B1
    t = ws["B1"]
    t.value     = f"📋  DETAIL TRANSAKSI — PERIODE {period_label.upper()}"
    t.font      = _font(bold=True, size=12, color=DARK_NAVY.lstrip("FF"))
    t.alignment = _align("left")

    # quick totals  B2/C2 and E2/F2
    ws["B2"].value = "Total Income:"
    ws["B2"].font  = _font(bold=True)
    ws["C2"].value = f'=SUMIF(C5:C{4+len(rows)},"Income",J5:J{4+len(rows)})'
    ws["C2"].number_format = RUPIAH_FMT
    ws["C2"].font  = _font(bold=True, color=GREEN_ACC.lstrip("FF"))

    ws["E2"].value = "Total Expense:"
    ws["E2"].font  = _font(bold=True)
    ws["F2"].value = f'=SUMIF(C5:C{4+len(rows)},"Expense",J5:J{4+len(rows)})'
    ws["F2"].number_format = RUPIAH_FMT
    ws["F2"].font  = _font(bold=True, color=RED_ACC.lstrip("FF"))

    # header row  row 4
    headers = ["Tanggal","Flow","Kategori","Sub Kategori","Keterangan",
               "Metode","Gross Amount (Rp)","Potongan (Rp)","Net Amount (Rp)"]
    for j, h in enumerate(headers):
        c = ws.cell(4, 2 + j)
        c.value     = h
        c.font      = _font(bold=True, size=10, color=WHITE_FONT)
        c.fill      = _fill(MID_NAVY)
        c.alignment = _align("center")
        c.border    = _border()
    ws.row_dimensions[4].height = 20

    # data rows
    for i, row in enumerate(rows):
        r   = 5 + i
        bg  = LIGHT_BG if i % 2 == 0 else WHITE
        is_income = row["Flow"] == "Income"

        vals = [
            row["Tanggal"],
            row["Flow"],
            row["Kategori"],
            row["Sub Kategori"],
            row["Keterangan"],
            row["Metode"],
            row["Gross Amount"],
            row["Potongan"],
            f"=H{r}-I{r}",   # Net Amount formula – keeps your original formula pattern
        ]

        for j, v in enumerate(vals):
            c = ws.cell(r, 2 + j)
            c.value  = v
            c.fill   = _fill(bg)
            c.border = _border()

            col_idx = j  # 0=Tanggal,1=Flow,...
            if col_idx == 1:  # Flow column
                c.font = _font(
                    bold=True,
                    color=RED_FONT.lstrip("FF") if not is_income else "FF27AE60"
                )
            elif col_idx in (6, 7):  # Gross, Potongan
                c.number_format = RUPIAH_FMT
            elif col_idx == 8:  # Net Amount
                c.number_format = RUPIAH_FMT

    # TOTAL row
    last_data = 4 + len(rows)
    tr = last_data + 1
    ws.cell(tr, 2).value = "TOTAL"
    ws.cell(tr, 2).font  = _font(bold=True)
    ws.cell(tr, 2).fill  = _fill(SECTION_BG)

    for j, col in [(6, "H"), (7, "I"), (8, "J")]:
        c = ws.cell(tr, 2 + j)
        c.value          = f"=SUM({col}5:{col}{last_data})"
        c.number_format  = RUPIAH_FMT
        c.font           = _font(bold=True)
        c.fill           = _fill(SECTION_BG)


# ── Ringkasan ─────────────────────────────────────────────────────────────────

def _build_ringkasan(wb, total_income, total_expense, net_profit, income_items, expense_items):
    ws = wb.create_sheet("Ringkasan")
    ws.sheet_view.showGridLines = False

    for col, w in [("A",2),("B",30),("C",20),("D",20)]:
        _set_col_width(ws, col, w)

    # summary header cells referenced by Dashboard  B2/C2/D2
    # B2=Total Pendapatan value, C2=Total Pengeluaran, D2=Laba Bersih
    ws["B2"].value         = total_income
    ws["B2"].number_format = RUPIAH_FMT
    ws["C2"].value         = total_expense
    ws["C2"].number_format = RUPIAH_FMT
    ws["D2"].value         = net_profit
    ws["D2"].number_format = RUPIAH_FMT

    # title  B1
    t = ws["B1"]
    t.value = "📑  RINGKASAN LAPORAN KEUANGAN"
    t.font  = _font(bold=True, size=12, color=DARK_NAVY.lstrip("FF"))

    # PENDAPATAN section
    ws["B3"].value = "PENDAPATAN"
    ws["B3"].font  = _font(bold=True, size=10, color=WHITE_FONT)
    ws["B3"].fill  = _fill(GREEN_ACC)

    start_income = 4
    for i, (label, val) in enumerate(income_items.items()):
        r = start_income + i
        ws.cell(r, 2).value = label
        ws.cell(r, 3).value = val
        ws.cell(r, 3).number_format = RUPIAH_FMT
        ws.cell(r, 3).fill = _fill(LIGHT_BG if i % 2 == 0 else WHITE)

    total_income_row = start_income + len(income_items)
    ws.cell(total_income_row, 2).value = "Total Pendapatan"
    ws.cell(total_income_row, 2).font  = _font(bold=True)
    ws.cell(total_income_row, 2).fill  = _fill(SECTION_BG)
    ws.cell(total_income_row, 3).value = f"=SUM(C{start_income}:C{total_income_row-1})"
    ws.cell(total_income_row, 3).number_format = RUPIAH_FMT
    ws.cell(total_income_row, 3).font  = _font(bold=True)
    ws.cell(total_income_row, 3).fill  = _fill(SECTION_BG)

    # PENGELUARAN section
    start_expense = total_income_row + 2
    ws.cell(start_expense - 1, 2).value = "PENGELUARAN"
    ws.cell(start_expense - 1, 2).font  = _font(bold=True, size=10, color=WHITE_FONT)
    ws.cell(start_expense - 1, 2).fill  = _fill(RED_ACC)

    for i, (label, val) in enumerate(expense_items.items()):
        r = start_expense + i
        ws.cell(r, 2).value = label
        ws.cell(r, 3).value = val
        ws.cell(r, 3).number_format = RUPIAH_FMT
        ws.cell(r, 3).fill = _fill(LIGHT_BG if i % 2 == 0 else WHITE)

    total_expense_row = start_expense + len(expense_items)
    ws.cell(total_expense_row, 2).value = "Total Pengeluaran"
    ws.cell(total_expense_row, 2).font  = _font(bold=True)
    ws.cell(total_expense_row, 2).fill  = _fill(SECTION_BG)
    ws.cell(total_expense_row, 3).value = f"=SUM(C{start_expense}:C{total_expense_row-1})"
    ws.cell(total_expense_row, 3).number_format = RUPIAH_FMT
    ws.cell(total_expense_row, 3).font  = _font(bold=True)
    ws.cell(total_expense_row, 3).fill  = _fill(SECTION_BG)

    # LABA / RUGI section
    laba_start = total_expense_row + 2
    ws.cell(laba_start, 2).value = "LABA / RUGI"
    ws.cell(laba_start, 2).font  = _font(bold=True, size=10, color=WHITE_FONT)
    ws.cell(laba_start, 2).fill  = _fill(BLUE_ACC)

    ws.cell(laba_start + 1, 2).value = "Laba Bersih"
    ws.cell(laba_start + 1, 2).font  = _font(bold=True)
    ws.cell(laba_start + 1, 3).value = (
        f"=C{total_income_row}-C{total_expense_row}"
    )
    ws.cell(laba_start + 1, 3).number_format = RUPIAH_FMT
    ws.cell(laba_start + 1, 3).font = _font(
        bold=True,
        color=GREEN_ACC.lstrip("FF") if net_profit >= 0 else RED_ACC.lstrip("FF")
    )

    # Dashboard references – overwrite B2/C2/D2 with formulas pointing to totals
    ws["B2"].value = f"=C{total_income_row}"
    ws["C2"].value = f"=C{total_expense_row}"
    ws["D2"].value = f"=C{laba_start+1}"


# ── Arus Kas ──────────────────────────────────────────────────────────────────

def _build_arus_kas(wb, period_label, rows):
    ws = wb.create_sheet("Arus Kas")
    ws.sheet_view.showGridLines = False

    for col, w in [("A",2),("B",16),("C",20),("D",20),("E",20),("F",2)]:
        _set_col_width(ws, col, w)

    # title
    t = ws["B1"]
    t.value = f"💵  LAPORAN ARUS KAS — {period_label.upper()}"
    t.font  = _font(bold=True, size=12, color=DARK_NAVY.lstrip("FF"))

    # header row 3
    for j, h in enumerate(["Tanggal", "Masuk (Rp)", "Keluar (Rp)", "Saldo (Rp)"]):
        c = ws.cell(3, 2 + j)
        c.value     = h
        c.font      = _font(bold=True, size=10, color=WHITE_FONT)
        c.fill      = _fill(MID_NAVY)
        c.alignment = _align("center")
        c.border    = _border()

    # data rows – same cumulative saldo formula pattern as template
    for i, row in enumerate(rows):
        r   = 4 + i
        bg  = LIGHT_BG if i % 2 == 0 else WHITE
        is_income = row["Flow"] == "Income"

        masuk  = row["Net Amount"] if is_income else 0
        keluar = row["Net Amount"] if not is_income else 0

        ws.cell(r, 2).value  = row["Tanggal"]
        ws.cell(r, 2).fill   = _fill(bg)
        ws.cell(r, 2).border = _border()

        mc = ws.cell(r, 3)
        mc.value         = masuk
        mc.number_format = RUPIAH_FMT
        mc.fill          = _fill(bg)
        mc.border        = _border()

        kc = ws.cell(r, 4)
        kc.value         = keluar
        kc.number_format = RUPIAH_FMT
        kc.fill          = _fill(bg)
        kc.border        = _border()

        sc = ws.cell(r, 5)
        if i == 0:
            sc.value = f"=C{r}-D{r}"
        else:
            sc.value = f"=E{r-1}+C{r}-D{r}"
        sc.number_format = RUPIAH_FMT
        sc.fill          = _fill(bg)
        sc.border        = _border()

    # TOTAL row
    last_data = 3 + len(rows)
    tr = last_data + 1
    ws.cell(tr, 2).value = "TOTAL"
    ws.cell(tr, 2).font  = _font(bold=True)
    ws.cell(tr, 2).fill  = _fill(SECTION_BG)

    tc = ws.cell(tr, 3)
    tc.value         = f"=SUM(C4:C{last_data})"
    tc.number_format = RUPIAH_FMT
    tc.font          = _font(bold=True)
    tc.fill          = _fill(SECTION_BG)

    dc = ws.cell(tr, 4)
    dc.value         = f"=SUM(D4:D{last_data})"
    dc.number_format = RUPIAH_FMT
    dc.font          = _font(bold=True)
    dc.fill          = _fill(SECTION_BG)

    sc = ws.cell(tr, 5)
    sc.value         = f"=C{tr}-D{tr}"
    sc.number_format = RUPIAH_FMT
    sc.font          = _font(bold=True)
    sc.fill          = _fill(SECTION_BG)


# ── quick test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    sample_data = [
        {"tanggal": "01/05/2026", "flow": "Expense", "kategori": "Pembelian",
         "sub_kategori": "Bahan Baku", "keterangan": "Polyflex", "metode": "Transfer",
         "gross_amount": 60500, "potongan": 0, "net_amount": 60500},
        {"tanggal": "01/05/2026", "flow": "Expense", "kategori": "Pembelian",
         "sub_kategori": "Bahan Baku", "keterangan": "Polyflex", "metode": "Transfer",
         "gross_amount": 398000, "potongan": 0, "net_amount": 398000},
        {"tanggal": "01/05/2026", "flow": "Expense", "kategori": "Pembelian",
         "sub_kategori": "Hangtag", "keterangan": "Hangtag", "metode": "Transfer",
         "gross_amount": 356000, "potongan": 0, "net_amount": 356000},
        {"tanggal": "01/05/2026", "flow": "Income", "kategori": "Income",
         "sub_kategori": "Modal", "keterangan": "Modal rendi", "metode": "Transfer",
         "gross_amount": 814500, "potongan": 0, "net_amount": 814500},
        {"tanggal": "14/05/2026", "flow": "Expense", "kategori": "Pembelian",
         "sub_kategori": "Lainnya", "keterangan": "Wallpaper", "metode": "Transfer",
         "gross_amount": 145000, "potongan": 0, "net_amount": 145000},
        {"tanggal": "03/05/2026", "flow": "Income", "kategori": "Income",
         "sub_kategori": "Penjualan", "keterangan": "Baju 1 pcs", "metode": "Cash",
         "gross_amount": 60000, "potongan": 0, "net_amount": 60000},
        {"tanggal": "03/05/2026", "flow": "Income", "kategori": "Income",
         "sub_kategori": "Penjualan", "keterangan": "Baju 1 pcs", "metode": "Transfer",
         "gross_amount": 50000, "potongan": 0, "net_amount": 50000},
    ]

    output = generate_excel_report(sample_data, "Mei", 2026)
    with open("/mnt/user-data/outputs/laporan_keuangan_kalren.xlsx", "wb") as f:
        f.write(output.read())
    print("Done.")